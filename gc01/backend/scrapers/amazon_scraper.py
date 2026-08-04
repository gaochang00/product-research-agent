"""
亚马逊产品数据爬虫 — 使用 Playwright 浏览器自动化
================================================
功能：自动搜索亚马逊指定类目，采集前50名产品数据（价格、评分、评论数、尺寸、功能等）
输出：Excel 文件，可直接导入分析系统

使用方法：
  1. 首次运行先安装依赖：python install_deps.py
  2. 运行爬虫：python amazon_scraper.py
  3. 爬取完成后，运行导入：python import_data.py

注意：脚本会打开Chrome浏览器窗口，请勿关闭，等待自动操作完成。
"""

import os
import re
import json
import time
import random
import logging
from datetime import datetime
from typing import List, Dict, Optional, Any
from pathlib import Path

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='[%(asctime)s] %(levelname)s: %(message)s',
    datefmt='%H:%M:%S'
)
logger = logging.getLogger(__name__)

# ============================================================
# 配置
# ============================================================
class Config:
    # 搜索关键词
    KEYWORD = "Clitoral Vibrators"
    
    # 目标采集数量
    TARGET_COUNT = 50
    
    # 采集详细信息和评论的产品数量（建议 30-50，越多越慢）
    DETAIL_TARGET_COUNT = 30
    
    # 每个产品最大评论采集页数（每页约10条评论）
    MAX_REVIEW_PAGES = 10
    
    # 是否打开浏览器窗口（True=可见，False=后台运行）
    HEADLESS = False
    
    # 输出文件路径
    OUTPUT_DIR = Path(__file__).parent.parent.parent / "data"
    OUTPUT_FILE = OUTPUT_DIR / f"amazon_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    OUTPUT_JSON = OUTPUT_DIR / f"amazon_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    
    # 亚马逊搜索URL
    AMAZON_URL = "https://www.amazon.com"
    
    # 操作延迟（秒）- 模拟真人操作，避免被检测
    MIN_DELAY = 1.5
    MAX_DELAY = 3.0


class AmazonScraper:
    """亚马逊产品数据爬虫"""

    def __init__(self):
        self.browser = None
        self.context = None
        self.page = None
        self.products = []
        self.collected_asins = set()
        self.playwright = None

    # ----------------------------------------------------------
    # 初始化浏览器
    # ----------------------------------------------------------
    def init_browser(self):
        """启动Chrome浏览器 - 使用Playwright独立浏览器实例
        
        使用专用用户数据目录，登录一次后自动保存会话状态。
        不会干扰您正在使用的Chrome浏览器。
        """
        from playwright.sync_api import sync_playwright
        
        # 专用用户数据目录（与您的个人Chrome完全隔离）
        scraper_data_dir = Config.OUTPUT_DIR.parent / "chrome_data"
        scraper_data_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"📁 使用专用用户数据目录: {scraper_data_dir}")
        
        logger.info("� 正在启动Chrome浏览器（独立实例，不干扰您的Chrome）...")
        
        self.playwright = sync_playwright().start()
        self.context = self.playwright.chromium.launch_persistent_context(
            user_data_dir=str(scraper_data_dir),
            channel="chrome",  # 使用系统已安装的Chrome
            headless=False,
            args=[
                '--disable-blink-features=AutomationControlled',
                '--no-first-run',
                '--no-default-browser-check',
                '--disable-sync',
                '--disable-features=ChromeWhatsNewUI',
                '--disable-background-timer-throttling',
                '--disable-renderer-backgrounding',
                '--disable-background-networking',
            ],
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
            locale="en-US",
            timezone_id="America/New_York",
            geolocation={"latitude": 40.7128, "longitude": -74.0060},
            permissions=["geolocation"],
        )
        
        # 获取当前页面
        self.page = self.context.pages[0] if self.context.pages else self.context.new_page()
        
        # 注入反检测脚本
        self.page.add_init_script("""
        Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
        Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
        Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
        """ )
        
        logger.info("✅ 浏览器启动完成")
        
        # 设置路由拦截（加速页面加载）
        self.page.route("**/*", self._block_unnecessary_requests)
        
        # ===== 引导用户登录（自动检测模式）=====
        logger.info("=" * 60)
        logger.info("请在刚打开的Chrome浏览器窗口中完成以下操作：")
        logger.info("")
        logger.info("  1. 访问 https://www.amazon.com")
        logger.info("  2. 点击右上角 'Sign in' 登录亚马逊账号")
        logger.info("     账号: 15306176642")
        logger.info("     密码: Q4wwVGsZzW2")
        logger.info("")
        logger.info("  3. 登录成功后，脚本会自动检测并继续")
        logger.info("     （自动检测中，请耐心等待...）")
        logger.info("=" * 60)
        
        # 先导航到Amazon首页
        try:
            self.page.goto("https://www.amazon.com", wait_until='domcontentloaded', timeout=60000)
            self._random_delay(2, 3)
        except:
            pass
        
        # 自动等待登录（最多180秒，每5秒检测一次）
        max_wait = 180  # 3分钟
        check_interval = 5
        logged_in = False
        
        for wait_sec in range(0, max_wait, check_interval):
            if wait_sec > 0:
                if wait_sec % 30 == 0:
                    logger.info(f"等待登录... ({wait_sec}秒 / {max_wait}秒)")
                time.sleep(check_interval)
            
            try:
                # 检查当前页面URL
                current_url = self.page.url
                
                # 如果不在登录页，尝试检测登录状态
                if "ap/signin" not in current_url and "signin" not in current_url.lower():
                    # 检测登录状态
                    try:
                        greeting = self.page.inner_text('#nav-link-accountList')
                        if greeting and ("Hello" in greeting):
                            logger.info(f"登录验证成功: {greeting.strip()[:80]}")
                            logged_in = True
                            break
                    except:
                        pass
                    
                    # 备用检测
                    if not logged_in:
                        try:
                            account_el = self.page.query_selector('#nav-link-accountList')
                            if account_el:
                                account_text = account_el.inner_text()
                                if account_text and "Sign in" not in account_text:
                                    logged_in = True
                                    logger.info("检测到账户信息，登录成功")
                                    break
                        except:
                            pass
                else:
                    # 仍在登录页，尝试刷新
                    if wait_sec > 0 and wait_sec % 30 == 0:
                        logger.info("仍在登录页面，请完成登录...")
            except:
                pass
        
        if not logged_in:
            logger.warning(f"等待 {max_wait} 秒后未检测到登录状态")
            logger.warning("请确认是否已登录，脚本将强制继续...")
        
        logger.info("浏览器初始化完成，开始采集数据...")

    def _block_unnecessary_requests(self, route):
        """屏蔽不必要的网络请求以加速加载"""
        url = route.request.url
        resource_type = route.request.resource_type
        
        # 允许文档和脚本请求
        if resource_type in ['document', 'script', 'xhr', 'fetch']:
            route.continue_()
        elif resource_type in ['image', 'media', 'font', 'stylesheet']:
            # 允许CSS和关键图片
            route.continue_()
        else:
            route.abort()

    # ----------------------------------------------------------
    # 搜索产品
    # ----------------------------------------------------------
    def search_products(self, keyword: str, skip_navigation: bool = False, auto_login: bool = True):
        """在亚马逊搜索关键词
        
        Args:
            keyword: 搜索关键词
            skip_navigation: 如果为True，不导航新页面，使用当前页面
            auto_login: 是否自动登录亚马逊
        """
        if skip_navigation:
            logger.info(f"🔍 使用现有页面，跳过导航")
            current_url = self.page.url
            current_title = self.page.title()
            logger.info(f"📍 当前页面URL: {current_url}")
            logger.info(f"📍 当前页面标题: {current_title}")
            self._random_delay(1, 2)
            self._handle_popups()
            return
        
        # 自动登录亚马逊
        if auto_login:
            self._amazon_login()
        
        logger.info(f"🔍 正在搜索: {keyword}")
        
        url = f"{Config.AMAZON_URL}/s?k={keyword.replace(' ', '+')}&ref=nb_sb_noss"
        self.page.goto(url, wait_until='domcontentloaded', timeout=90000)
        
        # 随机等待，模拟人类
        self._random_delay(3, 5)
        
        # 处理可能的验证码或Cookie弹窗
        self._handle_popups()
        
        # 调试：打印当前页面URL和标题
        current_url = self.page.url
        current_title = self.page.title()
        logger.info(f"📍 当前页面URL: {current_url}")
        logger.info(f"📍 当前页面标题: {current_title}")
        
        # 检查是否被重定向到验证码页面
        is_blocked = ("captcha" in current_url.lower() or "sorry" in current_url.lower() 
                      or "Sorry" in current_title or "captcha" in current_title.lower()
                      or "robot" in current_title.lower())
        
        if is_blocked:
            logger.warning("⚠️ 检测到亚马逊验证码/拦截页面！")
            logger.warning("   请在浏览器窗口中手动完成以下操作之一：")
            logger.warning("   1. 如果出现验证码，请手动勾选验证")
            logger.warning("   2. 如果出现拦截页面，请手动刷新或稍后重试")
            logger.warning("   3. 如果页面正常显示了搜索结果，请直接按回车继续")
            logger.warning("   ⏳ 等待你手动操作...")
            # 等待用户手动处理后继续
            try:
                input("   按回车键继续...")
            except:
                # 如果没有交互终端，等待30秒让用户手动操作
                logger.warning("   等待30秒让用户手动操作...")
                time.sleep(30)
            self._random_delay(2, 3)
            # 重新检查页面状态
            current_url = self.page.url
            current_title = self.page.title()
            logger.info(f"📍 当前页面URL: {current_url}")
            logger.info(f"📍 当前页面标题: {current_title}")
            # 如果仍然被拦截，尝试重新导航
            if ("captcha" in current_url.lower() or "sorry" in current_url.lower() 
                or "Sorry" in current_title or "captcha" in current_title.lower()):
                logger.warning("⚠️ 拦截仍然存在，尝试重新导航到搜索页...")
                for retry in range(3):
                    self._random_delay(5, 10)
                    self.page.goto(url, wait_until='domcontentloaded', timeout=60000)
                    self._random_delay(3, 5)
                    current_url = self.page.url
                    current_title = self.page.title()
                    logger.info(f"  重试{retry+1}: {current_title}")
                    if not ("captcha" in current_url.lower() or "sorry" in current_url.lower() 
                            or "Sorry" in current_title or "captcha" in current_title.lower()):
                        logger.info("✅ 拦截已解除！")
                        break
        
        # 检查是否跳转到中文站
        if "amazon.cn" in current_url or "amazon.com.cn" in current_url:
            logger.warning("⚠️ 被重定向到亚马逊中国站，尝试切换到.com")
            # 尝试添加语言参数
            self.page.goto(url + "&language=en_US", wait_until='domcontentloaded', timeout=60000)
            self._random_delay(2, 3)
        
        logger.info("✅ 搜索页面加载完成")

    def _handle_popups(self):
        """处理亚马逊可能出现的弹窗"""
        try:
            # Cookie 同意弹窗
            cookie_btn = self.page.query_selector('input[name="accept"]')
            if cookie_btn:
                cookie_btn.click()
                self._random_delay(0.5, 1)
        except:
            pass

    def _amazon_login(self):
        """自动登录亚马逊账号"""
        email = "15306176642"
        password = "Q4wwVGsZzW2"
        
        # 先获取当前页面信息
        try:
            current_url = self.page.url
            page_title = self.page.title()
        except:
            current_url = ""
            page_title = ""
        
        # 检查是否已经登录（如果当前不是登录页，说明已登录）
        if current_url and "ap/signin" not in current_url and "signin" not in current_url.lower():
            logger.info(f"✅ 已登录亚马逊（当前页面: {page_title}）")
            return
        
        # 尝试多个登录URL
        login_urls = [
            "https://www.amazon.com/ap/signin",
            "https://www.amazon.com/gp/sign-in.html",
        ]
        
        page_loaded = False
        for url in login_urls:
            try:
                logger.info(f"🌐 尝试访问登录页: {url}")
                self.page.goto(url, wait_until='domcontentloaded', timeout=30000)
                self._random_delay(2, 3)
                self._handle_popups()
                
                # 检查页面是否正常加载（不是404）
                try:
                    body_text = self.page.inner_text('body')[:200].lower()
                    if "looking for something" not in body_text and "sorry" not in body_text:
                        page_loaded = True
                        logger.info(f"✅ 登录页加载成功")
                        break
                    else:
                        logger.warning(f"  ⚠️ 该URL返回错误页面，尝试下一个")
                except:
                    page_loaded = True
                    break
            except Exception as e:
                logger.warning(f"  ⚠️ 访问登录页失败: {e}")
                continue
        
        if not page_loaded:
            # 所有登录URL都失败，直接导航到Amazon首页
            logger.info("🌐 导航到Amazon首页，尝试通过首页登录...")
            self.page.goto("https://www.amazon.com", wait_until='domcontentloaded', timeout=30000)
            self._random_delay(2, 3)
            self._handle_popups()
            
            # 检查是否已登录（首页会显示用户名）
            try:
                greeting = self.page.inner_text('#nav-link-accountList')
                if greeting and "Hello" in greeting:
                    logger.info(f"✅ 已登录亚马逊: {greeting.strip()}")
                    return
            except:
                pass
            
            # 点击"Sign in"按钮
            try:
                signin_btn = self.page.query_selector('#nav-link-accountList')
                if signin_btn:
                    signin_btn.click()
                    self._random_delay(2, 3)
                    try:
                        self.page.wait_for_load_state('domcontentloaded', timeout=15000)
                    except:
                        pass
                else:
                    self.page.goto("https://www.amazon.com/ap/signin?ie=UTF8", wait_until='domcontentloaded', timeout=30000)
                    self._random_delay(2, 3)
            except:
                pass
        
        # 检查是否已登录（页面跳转到了非登录页）
        try:
            current_url = self.page.url
            page_title = self.page.title()
        except:
            current_url = ""
            page_title = ""
        
        if "ap/signin" not in current_url and "signin" not in current_url.lower():
            logger.info(f"✅ 已登录亚马逊（当前页面: {page_title}）")
            return
        
        logger.info(f"📄 当前页面标题: {page_title}")
        
        # 检查页面内容，判断当前处于哪个登录步骤
        try:
            page_text = self.page.inner_text('body')[:500].lower()
        except:
            page_text = ""
        
        logger.info(f"📄 页面内容片段: {page_text[:150]}")
        
        # 情况1: 直接显示"继续为[账号名]"按钮（已记住账号）
        try:
            continue_name_btn = self.page.query_selector('input[data-a-input-name*="rememberMe"]')
            if continue_name_btn:
                logger.info("  检测到'继续为[账号名]'按钮，点击继续...")
                continue_name_btn.click()
                self._random_delay(2, 3)
                try:
                    self.page.wait_for_load_state('domcontentloaded', timeout=15000)
                except:
                    pass
                return self._fill_password_and_login(password)
        except:
            pass
        
        # 情况2: 显示邮箱输入框
        logger.info("📧 正在查找邮箱输入框...")
        email_input = None
        email_selectors = [
            '#ap_email', 'input[type="email"]', 'input[name="email"]', 
            'input.a-input-text', 'input[aria-label*="Email"]', 'input[aria-label*="email"]',
            'input[type="text"]'
        ]
        for sel in email_selectors:
            try:
                el = self.page.wait_for_selector(sel, timeout=3000)
                if el and el.is_visible():
                    email_input = el
                    logger.info(f"  找到邮箱输入框: {sel}")
                    break
            except:
                continue
        
        if email_input:
            try:
                email_input.click()
                self._random_delay(0.5, 1)
                email_input.fill(email)
                self._random_delay(1, 2)
                
                continue_btn = self.page.query_selector('#continue')
                if continue_btn:
                    continue_btn.click()
                else:
                    email_input.press('Enter')
                self._random_delay(2, 3)
                try:
                    self.page.wait_for_load_state('domcontentloaded', timeout=15000)
                except:
                    pass
                
                return self._fill_password_and_login(password)
            except Exception as e:
                logger.warning(f"  邮箱输入异常: {e}")
        
        # 情况3: 直接显示密码输入框（已记住账号）
        logger.info("  尝试直接输入密码...")
        try:
            pw_input = self.page.wait_for_selector('#ap_password', timeout=5000)
            if pw_input and pw_input.is_visible():
                logger.info("  找到密码输入框，直接输入密码...")
                pw_input.click()
                self._random_delay(0.5, 1)
                pw_input.fill(password)
                self._random_delay(1, 2)
                signin_btn = self.page.query_selector('#signInSubmit')
                if signin_btn:
                    signin_btn.click()
                    logger.info("⏳ 正在登录...")
                    self._random_delay(3, 5)
                    self._wait_login_complete()
                    return
                else:
                    pw_input.press('Enter')
                    self._random_delay(3, 5)
                    self._wait_login_complete()
                    return
        except:
            pass
        
        # 情况4: 尝试用更通用的方式查找密码输入
        try:
            pw_input = self.page.query_selector('input[type="password"]')
            if pw_input:
                logger.info("  找到通用密码输入框...")
                pw_input.fill(password)
                self._random_delay(1, 2)
                submit_btn = self.page.query_selector('#signInSubmit, input[type="submit"], button[type="submit"]')
                if submit_btn:
                    submit_btn.click()
                    self._random_delay(3, 5)
                    self._wait_login_complete()
                    return
        except:
            pass
        
        # 所有自动方法都失败，回退到手动登录
        logger.warning("⚠️ 自动登录无法完成，请手动登录")
        self._wait_manual_login()
    
    def _fill_password_and_login(self, password: str):
        """输入密码并点击登录（在已输入邮箱后调用）"""
        logger.info("🔑 正在输入密码...")
        try:
            pw_input = self.page.wait_for_selector('#ap_password', timeout=15000)
            if pw_input and pw_input.is_visible():
                pw_input.click()
                self._random_delay(0.5, 1)
                pw_input.fill(password)
                self._random_delay(1, 2)
                
                signin_btn = self.page.query_selector('#signInSubmit')
                if signin_btn:
                    signin_btn.click()
                    logger.info("⏳ 正在登录...")
                    self._random_delay(3, 5)
                    self._wait_login_complete()
                    return
                else:
                    pw_input.press('Enter')
                    self._random_delay(3, 5)
                    self._wait_login_complete()
                    return
            else:
                logger.warning("  密码输入框未找到或不可见")
        except Exception as e:
            logger.warning(f"  密码输入异常: {e}")
        
        # 尝试通用密码输入
        try:
            pw_input = self.page.query_selector('input[type="password"]')
            if pw_input:
                logger.info("  使用通用密码输入框...")
                pw_input.fill(password)
                self._random_delay(1, 2)
                submit_btn = self.page.query_selector('#signInSubmit, input[type="submit"], button[type="submit"]')
                if submit_btn:
                    submit_btn.click()
                    self._random_delay(3, 5)
                    self._wait_login_complete()
                    return
        except:
            pass
        
        logger.warning("⚠️ 自动输入密码失败，请手动登录")
        self._wait_manual_login()
    
    def _wait_login_complete(self):
        """等待登录完成（自动检测或等待OTP）"""
        for wait_sec in range(120):
            time.sleep(1)
            try:
                current_url = self.page.url
                page_text = self.page.inner_text('body')[:300].lower()
            except:
                continue
            
            # 检查是否登录成功
            if "ap/signin" not in current_url and "signin" not in current_url.lower() and "signin" not in page_text:
                logger.info(f"✅ 亚马逊登录成功")
                self._random_delay(2, 3)
                return
            
            # 检查是否需要OTP验证码
            if any(kw in page_text for kw in ["otp", "verification code", "mfa", "authenticator", "enter code", "one-time password"]):
                logger.warning("⚠️ 需要OTP验证码（短信/邮箱验证）")
                logger.warning("   请在Chrome浏览器中查看验证码输入框")
                logger.warning("   输入验证码后，脚本会自动继续")
                for otp_sec in range(120):
                    time.sleep(1)
                    try:
                        current_url = self.page.url
                        if "ap/signin" not in current_url and "signin" not in current_url.lower():
                            logger.info(f"✅ 验证码验证成功")
                            self._random_delay(2, 3)
                            return
                    except:
                        pass
                    if otp_sec % 30 == 0 and otp_sec > 0:
                        logger.warning(f"   等待验证码输入... ({otp_sec}秒)")
                break
            
            if wait_sec % 20 == 0 and wait_sec > 0:
                logger.info(f"   等待登录完成... ({wait_sec}秒)")
        
        logger.warning("  登录超时")
    
    def _wait_manual_login(self):
        """等待用户手动登录"""
        logger.warning("   请在Chrome浏览器中完成登录，然后按回车键继续")
        try:
            input()
        except:
            for wait_sec in range(120):
                time.sleep(1)
                try:
                    current_url = self.page.url
                    if "ap/signin" not in current_url and "signin" not in current_url.lower():
                        logger.info(f"✅ 手动登录成功")
                        return
                except:
                    pass

    # ----------------------------------------------------------
    # 采集产品列表
    # ----------------------------------------------------------
    def collect_product_list(self):
        """从搜索结果页采集产品列表"""
        logger.info("📋 开始采集产品列表...")
        
        # 多尝试几种选择器，适应亚马逊不同的页面结构
        selectors = [
            'div[data-component-type="s-search-result"]',
            'div[data-asin]:not([data-asin=""])',
            '.s-result-item',
            'div[cel_widget_id^="MAIN-SEARCH_RESULTS"]',
        ]
        
        # 先保存页面截图用于调试
        try:
            screenshot_path = Config.OUTPUT_DIR / "debug_screenshot.png"
            self.page.screenshot(path=str(screenshot_path))
            logger.info(f"📸 已保存页面截图: {screenshot_path}")
        except:
            pass
        
        # 打印页面HTML的前2000字符用于调试
        try:
            body_text = self.page.inner_text('body')[:500]
            logger.info(f"📝 页面文本预览: {body_text[:200]}")
        except:
            pass
        
        scroll_count = 0
        max_scrolls = 20
        no_new_count = 0
        
        # 查找有效的选择器
        found_selector = None
        for sel in selectors:
            count = self.page.query_selector_all(sel).__len__() if hasattr(self.page.query_selector_all(sel), '__len__') else len(self.page.query_selector_all(sel))
            try:
                count = len(self.page.query_selector_all(sel))
                if count > 0:
                    found_selector = sel
                    logger.info(f"  使用选择器 '{sel}' 找到 {count} 个产品")
                    break
            except:
                continue
        
        if not found_selector:
            logger.warning("⚠️ 所有选择器都无法匹配，可能页面结构异常")
            logger.warning("   尝试使用备用方案: 滚动页面后重试...")
            # 尝试滚动页面
            self.page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            self._random_delay(2, 3)
            for sel in selectors:
                try:
                    count = len(self.page.query_selector_all(sel))
                    if count > 0:
                        found_selector = sel
                        logger.info(f"  滚动后使用选择器 '{sel}' 找到 {count} 个产品")
                        break
                except:
                    continue
        
        if not found_selector:
            logger.error("❌ 无法找到产品列表，请检查浏览器中显示的页面")
            return
        
        while len(self.collected_asins) < Config.TARGET_COUNT and scroll_count < max_scrolls:
            scroll_count += 1
            
            # 等待产品卡片加载
            try:
                self.page.wait_for_selector(found_selector, timeout=15000)
            except:
                logger.warning("⚠️ 等待超时")
                break
            
            # 提取当前页所有产品
            items = self.page.query_selector_all(found_selector)
            logger.info(f"  第{scroll_count}页: 找到 {len(items)} 个产品")
            
            # 调试：打印第一个产品的HTML结构
            if scroll_count == 1 and items:
                try:
                    first_html = items[0].inner_html()[:500]
                    logger.info(f"  📝 第一个产品HTML结构: {first_html[:200]}")
                    first_asin = items[0].get_attribute('data-asin')
                    logger.info(f"  📝 data-asin属性: {first_asin}")
                except Exception as e:
                    logger.info(f"  📝 无法获取HTML: {e}")
            
            new_count = 0
            for item in items:
                try:
                    # 尝试多种方式获取ASIN
                    asin = item.get_attribute('data-asin')
                    if not asin:
                        # 尝试从链接中提取ASIN
                        link = item.query_selector('a[href*="/dp/"]')
                        if link:
                            href = link.get_attribute('href')
                            asin_match = re.search(r'/dp/([A-Z0-9]{10})', href)
                            if asin_match:
                                asin = asin_match.group(1)
                    if not asin or asin in self.collected_asins:
                        continue
                    
                    product = self._extract_list_item(item, asin)
                    if product:
                        self.products.append(product)
                        self.collected_asins.add(asin)
                        new_count += 1
                        
                except Exception as e:
                    logger.debug(f"  提取产品失败: {e}")
            
            if new_count == 0:
                no_new_count += 1
                if no_new_count >= 3:
                    logger.info("  连续3页无新数据，停止采集")
                    break
            else:
                no_new_count = 0
            
            logger.info(f"  ✅ 已采集 {len(self.collected_asins)}/{Config.TARGET_COUNT} 个产品")
            
            # 如果够了就停止
            if len(self.collected_asins) >= Config.TARGET_COUNT:
                break
            
            # 翻到下一页
            if not self._go_to_next_page():
                logger.info("📄 没有更多页面了")
                break
        
        logger.info(f"📊 列表采集完成: 共 {len(self.products)} 个产品")

    def _extract_list_item(self, item, asin: str) -> Optional[Dict]:
        """从搜索结果卡片中提取产品信息"""
        try:
            # 调试：打印第一个产品的部分HTML
            if asin == list(self.collected_asins)[0] if self.collected_asins else True:
                pass  # 只在第一页第一个产品调试
            
            # 标题 - 尝试多种选择器
            title = ''
            title_selectors = [
                'h2 a span',
                'h2 span.a-text-normal',
                'h2 a.a-link-normal span.a-text-normal',
                'a.a-link-normal span.a-text-normal',
                '[data-cy="title-recipe"] a',
                'h2 a',
            ]
            for sel in title_selectors:
                try:
                    el = item.query_selector(sel)
                    if el:
                        t = el.inner_text().strip()
                        if t and len(t) > 5:
                            title = t
                            break
                except:
                    continue
            
            if not title:
                # 最后的尝试：从任何h2中获取文本
                try:
                    h2 = item.query_selector('h2')
                    if h2:
                        title = h2.inner_text().strip()
                except:
                    pass
            
            if not title:
                return None
            
            # 价格
            price = 0
            try:
                price_el = item.query_selector('span.a-price[data-a-size^="xl"] span.a-offscreen, span.a-price[data-a-size^="l"] span.a-offscreen, span.a-price span.a-offscreen')
                if not price_el:
                    price_el = item.query_selector('span.a-price-whole')
                if price_el:
                    price_text = price_el.inner_text().strip().replace('$', '').replace(',', '')
                    price = float(price_text)
            except:
                try:
                    price_el = item.query_selector('span.a-price')
                    if price_el:
                        offscreen = price_el.query_selector('span.a-offscreen')
                        if offscreen:
                            price_text = offscreen.inner_text().strip().replace('$', '').replace(',', '')
                            price = float(price_text)
                except:
                    pass
                    price = 0
            
            # 评分
            rating_el = item.query_selector('i.a-icon-star-small span.a-icon-alt')
            rating = 0
            if rating_el:
                rating_text = rating_el.inner_text().strip()
                rating_match = re.search(r'([\d.]+)', rating_text)
                if rating_match:
                    rating = float(rating_match.group(1))
            
            # 评论数（多种选择器）
            review_count = 0
            review_selectors = [
                'span.a-size-base.s-underline-text',
                'a.a-link-normal span.a-size-base',
                'span.a-size-base[data-hook="review-count"]',
                'a[href*="customer-reviews"] span.a-size-base',
            ]
            for sel in review_selectors:
                try:
                    review_el = item.query_selector(sel)
                    if review_el:
                        review_text = review_el.inner_text().strip().replace(',', '').replace('(', '').replace(')', '')
                        # 提取数字
                        nums = re.findall(r'\d+', review_text)
                        if nums:
                            review_count = int(nums[0])
                            break
                except:
                    continue
            
            # 图片URL
            img_el = item.query_selector('img.s-image')
            main_image = img_el.get_attribute('src') if img_el else ''
            
            # 品牌
            brand_el = item.query_selector('h5[data-attribute]')
            brand = brand_el.get_attribute('data-attribute') if brand_el else ''
            if not brand:
                # 从标题中提取品牌（通常是第一个词）
                brand = title.split(' ')[0] if title else ''
            
            product_url = f"https://www.amazon.com/dp/{asin}"
            
            return {
                "asin": asin,
                "title": title,
                "brand": brand,
                "price": price,
                "rating": rating,
                "review_count": review_count,
                "main_image": main_image,
                "product_url": product_url,
                "source": "search_list",
                "bullet_points": [],
                "specifications": {},
                "features": [],
                "color_options": [],
                "material": "",
                "dimensions": "",
                "weight": "",
                "description": "",
                "bsr_category": ""
            }
        except Exception as e:
            logger.debug(f"  提取列表项失败: {e}")
            return None

    def _go_to_next_page(self) -> bool:
        """翻到下一页"""
        try:
            next_btn = self.page.query_selector('a.s-pagination-next')
            if not next_btn:
                return False
            
            # 检查是否被禁用
            if next_btn.get_attribute('aria-disabled') == 'true':
                return False
            
            # 滚动到翻页按钮
            next_btn.scroll_into_view_if_needed()
            self._random_delay(0.5, 1)
            
            next_btn.click()
            self.page.wait_for_load_state('networkidle', timeout=30000)
            self._random_delay(2, 4)
            
            return True
        except Exception as e:
            logger.warning(f"⚠️ 翻页失败: {e}")
            return False

    # ----------------------------------------------------------
    # 采集产品详情
    # ----------------------------------------------------------
    def _ensure_review_login(self):
        """确保评论页的登录状态，在批量采集评论前先完成登录"""
        logger.info("  🔑 正在检查评论页登录状态...")
        try:
            if self.products:
                first_asin = self.products[0]["asin"]
                reviews_url = f"https://www.amazon.com/product-reviews/{first_asin}"
                self.page.goto(reviews_url, wait_until='domcontentloaded', timeout=60000)
                self._random_delay(2, 3)
                self._handle_popups()
                
                current_url = self.page.url
                if "ap/signin" in current_url or "signin" in current_url.lower():
                    logger.warning("  ⚠️ 评论页需要登录，尝试自动登录...")
                    # 尝试使用自动登录
                    self._amazon_login()
                    # 重新访问评论页
                    self.page.goto(reviews_url, wait_until='domcontentloaded', timeout=60000)
                    self._random_delay(2, 3)
                    current_url = self.page.url
                    if "ap/signin" not in current_url and "signin" not in current_url.lower():
                        logger.info("  ✅ 评论页登录成功")
                    else:
                        logger.warning("  ⚠️ 评论页登录失败，将跳过评论采集")
                else:
                    logger.info("  ✅ 评论页已成功登录")
        except Exception as e:
            logger.warning(f"  ⚠️ 检查评论页登录状态失败: {e}")

    def collect_product_details(self):
        """进入每个产品的详情页，采集更多数据"""
        logger.info("🔍 开始采集产品详情...")
        
        # 采集前 DETAIL_TARGET_COUNT 个产品的详情和评论
        detail_targets = self.products[:Config.DETAIL_TARGET_COUNT]
        
        # 先尝试登录评论页，确保后续评论采集可以正常进行
        self._ensure_review_login()
        
        for i, product in enumerate(detail_targets):
            asin = product["asin"]
            title = product["title"][:30]
            logger.info(f"  [{i+1}/{len(detail_targets)}] 采集详情: {title}...")
            
            try:
                self._scrape_detail_page(product)
                logger.info(f"    ✅ 完成")
            except Exception as e:
                logger.warning(f"    ⚠️ 采集详情失败: {e}")
            
            # 随机延迟
            self._random_delay(2, 4)

    def _scrape_detail_page(self, product: Dict):
        """采集单个产品详情页"""
        asin = product["asin"]
        url = f"https://www.amazon.com/dp/{asin}"
        
        # 打开详情页
        try:
            self.page.goto(reviews_url, wait_until='domcontentloaded', timeout=60000)
        except:
            try:
                self.page.goto(url, wait_until='load', timeout=60000)
            except:
                logger.warning(f"    ⚠️ 详情页加载超时，尝试继续...")
        self._random_delay(1, 2)
        
        # 处理弹窗
        self._handle_popups()
        
        # --- 提取 bullet points（产品卖点） ---
        try:
            bullet_els = self.page.query_selector_all('#feature-bullets li span.a-list-item')
            product["bullet_points"] = [
                el.inner_text().strip() 
                for el in bullet_els 
                if el.inner_text().strip()
            ]
        except:
            pass
        
        # --- 提取产品描述 ---
        try:
            desc_el = self.page.query_selector('#productDescription p')
            if desc_el:
                product["description"] = desc_el.inner_text().strip()[:500]
        except:
            pass
        
        # --- 提取规格表 ---
        try:
            spec_rows = self.page.query_selector_all('#productDetails_techSpec_section_1 tr')
            for row in spec_rows:
                cells = row.query_selector_all('th, td')
                if len(cells) >= 2:
                    key = cells[0].inner_text().strip().rstrip(':')
                    val = cells[1].inner_text().strip()
                    product["specifications"][key] = val
                    
                    # 提取关键字段
                    if '尺寸' in key or 'dimension' in key.lower():
                        product["dimensions"] = val
                    elif '重量' in key or 'weight' in key.lower():
                        product["weight"] = val
                    elif 'material' in key.lower() or '材质' in key:
                        product["material"] = val
        except:
            pass
        
        # --- 提取颜色选项 ---
        try:
            color_els = self.page.query_selector_all('#variation_color_name li')
            product["color_options"] = []
            for el in color_els:
                img = el.query_selector('img')
                if img:
                    alt = img.get_attribute('alt')
                    if alt and alt not in product["color_options"]:
                        product["color_options"].append(alt)
        except:
            pass
        
        # --- 提取BSR排名 ---
        try:
            bsr_el = self.page.query_selector('#productDetails_detailBullets_sections1 tr:has(th:has-text("Best Sellers Rank")) td')
            if bsr_el:
                product["bsr_category"] = bsr_el.inner_text().strip()[:100]
        except:
            pass
        
        # --- 从标题和卖点中提取功能关键词 ---
        features = set()
        text_to_search = product["title"] + " " + " ".join(product["bullet_points"])
        text_lower = text_to_search.lower()
        
        feature_keywords = {
            "空气脉冲": ["air pulse", "pleasure air"],
            "声波": ["sonic", "sound wave"],
            "APP控制": ["app", "bluetooth", "remote", "smart"],
            "防水": ["waterproof", "ipx", "shower"],
            "静音": ["quiet", "silent", "whisper"],
            "USB充电": ["usb", "rechargeable"],
            "无线充电": ["wireless", "qi"],
            "多种模式": ["mode", "pattern", "speed", "intensity"],
            "人体工学": ["ergonomic", "curved"],
            "医用硅胶": ["medical", "silicone", "body-safe"],
            "旅行锁": ["travel lock", "safety lock"],
            "情侣使用": ["couple", "partner", "wearable"],
        }
        
        for feat_name, keywords in feature_keywords.items():
            if any(k in text_lower for k in keywords):
                features.add(feat_name)
        
        product["features"] = list(features)
        
        # --- 采集评论 ---
        self._scrape_reviews(product, max_pages=Config.MAX_REVIEW_PAGES)

    def _scrape_reviews(self, product: Dict, max_pages: int = None):
        """采集产品评论数据 - 从产品详情页点击"查看评论"进入评论页
        
        Args:
            product: 产品字典（会添加 reviews 字段）
            max_pages: 最多翻页数，每页约10条评论，默认 Config.MAX_REVIEW_PAGES
        """
        if max_pages is None:
            max_pages = Config.MAX_REVIEW_PAGES
        asin = product["asin"]
        
        product["reviews"] = []
        product["review_summary"] = {}
        
        logger.info(f"    💬 正在采集评论 ({asin})...")
        
        # 评论页URL（提前定义，供多个位置使用）
        reviews_url = f"https://www.amazon.com/product-reviews/{asin}"
        
        try:
            # 先从产品详情页点击"See all reviews"进入评论页
            # 当前页面应该在产品详情页（_scrape_detail_page 已加载）
            
            # 尝试找到"See all reviews"链接并点击
            review_link = None
            review_selectors = [
                'a[data-hook="see-all-reviews-link-foot"]',
                'a[data-hook="see-all-rating-link"]',
                'a.a-link-emphasis[href*="product-reviews"]',
                'a[href*="product-reviews"]',
            ]
            
            for sel in review_selectors:
                try:
                    el = self.page.query_selector(sel)
                    if el and el.is_visible():
                        review_link = el
                        break
                except:
                    continue
            
            if review_link:
                logger.info(f"      � 点击'查看评论'链接")
                review_link.click()
                self._random_delay(2, 4)
                try:
                    self.page.wait_for_load_state('domcontentloaded', timeout=30000)
                except:
                    pass
            else:
                # 如果找不到链接，直接导航到评论页
                self.page.goto(reviews_url, wait_until='domcontentloaded', timeout=60000)
                self._random_delay(3, 5)
                self._handle_popups()
                
                current_url = self.page.url
                if "/product-reviews/" not in current_url:
                    logger.info(f"      ⏭️ 无法访问评论页（被重定向）")
                    return
            
            # 检查是否被重定向到登录页
            current_url = self.page.url
            if "ap/signin" in current_url or "signin" in current_url.lower():
                logger.warning(f"      ⚠️ 评论页需要登录，尝试自动登录...")
                self._amazon_login()
                # 重新访问评论页
                self.page.goto(reviews_url, wait_until='domcontentloaded', timeout=60000)
                self._random_delay(2, 3)
                current_url = self.page.url
                if "ap/signin" in current_url or "signin" in current_url.lower():
                    logger.info(f"      ⏭️ 自动登录失败，跳过该产品评论")
                    return
                else:
                    logger.info(f"      ✅ 自动登录成功！继续采集评论")
            
            # 检查页面是否显示"no reviews"
            try:
                page_text = self.page.inner_text('body')
                if "no customer reviews" in page_text.lower() or "there are no reviews" in page_text.lower():
                    logger.info(f"      ℹ️ 该产品暂无评论")
                    return
            except:
                pass
            
            # 等待评论区域加载
            try:
                self.page.wait_for_selector('#cm_cr-review_list', timeout=20000)
                self._random_delay(1, 2)
            except:
                pass
            
            # 尝试按"最近"排序
            try:
                self.page.evaluate('''() => {
                    const links = document.querySelectorAll('a[href*="sortBy=recent"]');
                    for (const link of links) {
                        if (link.offsetParent !== null) {
                            link.click();
                            return;
                        }
                    }
                }''')
                self._random_delay(2, 3)
                try:
                    self.page.wait_for_load_state('domcontentloaded', timeout=15000)
                except:
                    pass
            except:
                pass
            
            seen_reviews = set()
            
            for page_num in range(max_pages):
                # 等待评论元素
                try:
                    self.page.wait_for_selector('[data-hook="review"]', timeout=20000)
                except:
                    logger.info(f"      ⏭️ 未找到评论（第{page_num+1}页）")
                    break
                
                review_elements = self.page.query_selector_all('[data-hook="review"]')
                
                new_count = 0
                for el in review_elements:
                    try:
                        review = {}
                        
                        rating_el = el.query_selector('[data-hook="review-star-rating"]')
                        if rating_el:
                            rating_text = rating_el.inner_text().strip()
                            rating_match = re.search(r'([\d.]+)', rating_text)
                            review["rating"] = float(rating_match.group(1)) if rating_match else 0
                        else:
                            review["rating"] = 0
                        
                        title_el = el.query_selector('[data-hook="review-title"]')
                        review["title"] = title_el.inner_text().strip() if title_el else ""
                        
                        # 先尝试展开被截断的评论（点击"Read more" / "展开"链接）
                        try:
                            # Amazon 截断评论文本，需要点击展开才能获取完整内容
                            expand_links = el.query_selector_all('a[data-hook="see-all-review-links"], a.a-link-normal[href*="read-more"], a[href*="read-more"], span.reread-link a, a[class*="expand"], .review-text a:has-text("more"), a:has-text("Read more"), a:has-text("more")')
                            for link in expand_links:
                                if link and link.is_visible():
                                    link.click()
                                    import time
                                    time.sleep(0.3)  # 等待展开动画
                                    break
                        except:
                            pass
                        # 备用: 使用JS展开所有评论
                        try:
                            self.page.evaluate('''() => {
                                document.querySelectorAll('[data-hook="review-body"] a').forEach(a => {
                                    if(a.offsetParent !== null) a.click();
                                });
                            }''')
                            import time
                            time.sleep(0.3)
                        except:
                            pass
                        
                        # 获取完整评论内容（展开后所有文本都被加载）
                        body_el = el.query_selector('[data-hook="review-body"]')
                        if body_el:
                            # 获取纯文本（跳过"Read more"等链接文本）
                            review["text"] = body_el.inner_text().strip()
                            # 去掉末尾可能的展开/收起操作文字
                            for suffix in ["Read more", "less", "展开", "收起", "more", "less"]:
                                if review["text"].endswith(suffix):
                                    review["text"] = review["text"][:-len(suffix)].strip()
                        else:
                            review["text"] = ""
                        
                        date_el = el.query_selector('[data-hook="review-date"]')
                        review["date"] = date_el.inner_text().strip() if date_el else ""
                        
                        helpful_el = el.query_selector('[data-hook="helpful-vote-statement"]')
                        review["helpful"] = helpful_el.inner_text().strip() if helpful_el else ""
                        
                        verified_el = el.query_selector('[data-hook="avp-badge"]')
                        review["verified"] = bool(verified_el)
                        
                        dedup_key = (review["title"] + review["text"][:50]).strip()
                        if dedup_key and dedup_key not in seen_reviews and len(review["text"]) > 10:
                            seen_reviews.add(dedup_key)
                            product["reviews"].append(review)
                            new_count += 1
                        
                    except:
                        continue
                
                logger.info(f"      📝 第{page_num+1}页: 采集 {new_count} 条评论")
                
                # 判断是否还有更多页
                if page_num + 1 >= max_pages:
                    break
                
                has_next = False
                
                # 策略1: 使用标准亚马逊评论翻页URL格式直接导航
                try:
                    next_page = page_num + 2
                    # 标准Amazon评论翻页URL格式
                    next_url = f"https://www.amazon.com/product-reviews/{asin}/ref=cm_cr_getr_d_paging_btm_next_{next_page}?ie=UTF8&reviewerType=all_reviews&pageNumber={next_page}"
                    self.page.goto(next_url, wait_until='domcontentloaded', timeout=30000)
                    self._random_delay(2, 4)
                    has_next = True
                except:
                    pass
                
                # 策略2: 点击分页栏中的下一页按钮
                if not has_next:
                    try:
                        pagination = self.page.query_selector('.a-pagination')
                        if pagination:
                            next_page_el = pagination.query_selector('li.a-last a')
                            if next_page_el:
                                href = next_page_el.get_attribute('href')
                                if href and href != '#':
                                    next_page_el.click()
                                    self._random_delay(2, 4)
                                    try:
                                        self.page.wait_for_load_state('domcontentloaded', timeout=20000)
                                    except:
                                        pass
                                    has_next = True
                    except:
                        pass
                
                if not has_next:
                    logger.info(f"      ⏹️ 没有更多评论页")
                    break
            
            if product["reviews"]:
                ratings = [r["rating"] for r in product["reviews"]]
                total = len(ratings)
                avg_rating = sum(ratings) / total if total else 0
                positive = len([r for r in product["reviews"] if r["rating"] >= 4])
                negative = len([r for r in product["reviews"] if r["rating"] <= 2])
                
                product["review_summary"] = {
                    "total_scraped": total,
                    "average_rating": round(avg_rating, 2),
                    "positive_count": positive,
                    "negative_count": negative,
                    "positive_ratio": round(positive / total * 100, 1) if total else 0
                }
                logger.info(f"      📊 评论摘要: 共{total}条, 均分{avg_rating:.1f}, 好评率{product['review_summary']['positive_ratio']}%")
            
        except Exception as e:
            logger.warning(f"    ⚠️ 采集评论失败: {e}")

    # ----------------------------------------------------------
    # 数据导出
    # ----------------------------------------------------------
    def export_to_json(self):
        """导出为JSON格式"""
        Config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        
        # 构建与演示数据兼容的格式
        export_data = []
        for p in self.products:
            export_data.append({
                "asin": p["asin"],
                "title": p["title"],
                "brand": p["brand"],
                "price": p["price"],
                "currency": "USD",
                "rating": p["rating"],
                "review_count": p["review_count"],
                "main_image": p["main_image"],
                "product_url": p["product_url"],
                "category": "女性情趣用品",
                "subcategory": Config.KEYWORD,
                "bsr_category": p["bsr_category"],
                "description": p["description"],
                "bullet_points": p["bullet_points"],
                "specifications": p["specifications"],
                "dimensions": p["dimensions"],
                "weight": p["weight"],
                "color_options": p["color_options"],
                "material": p["material"],
                "features": p["features"],
                "estimated_monthly_sales": 0,
                # 评论数据
                "reviews": p.get("reviews", []),
                "review_summary": p.get("review_summary", {})
            })
        
        # 写入JSON
        with open(Config.OUTPUT_JSON, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"📄 JSON 已导出: {Config.OUTPUT_JSON}")
        return Config.OUTPUT_JSON

    def export_to_excel(self):
        """导出为Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            Config.OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "亚马逊产品数据"
            
            # 表头
            headers = [
                "ASIN", "标题", "品牌", "价格(USD)", "评分", "评论数",
                "图片URL", "产品链接", "子品类", "BSR排名",
                "尺寸", "重量", "材质", "颜色数", "颜色选项",
                "功能特点", "描述"
            ]
            
            # 样式
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="7C5CFC", end_color="7C5CFC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # 数据行
            for row_idx, p in enumerate(self.products, 2):
                data = [
                    p["asin"],
                    p["title"],
                    p["brand"],
                    p["price"],
                    p["rating"],
                    p["review_count"],
                    p["main_image"],
                    p["product_url"],
                    Config.KEYWORD,
                    p["bsr_category"],
                    p["dimensions"],
                    p["weight"],
                    p["material"],
                    len(p["color_options"]),
                    ", ".join(p["color_options"]),
                    ", ".join(p["features"]),
                    p["description"][:200] if p["description"] else ""
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=(col == 2))
            
            # 设置列宽
            col_widths = [15, 50, 15, 12, 8, 10, 40, 40, 20, 30, 15, 10, 20, 8, 25, 30, 40]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = width
            
            wb.save(str(Config.OUTPUT_FILE))
            logger.info(f"📄 Excel 已导出: {Config.OUTPUT_FILE}")
            return Config.OUTPUT_FILE
            
        except ImportError:
            logger.warning("⚠️ openpyxl 未安装，跳过Excel导出")
            return None

    # ----------------------------------------------------------
    # 工具方法
    # ----------------------------------------------------------
    def _random_delay(self, min_sec: float = None, max_sec: float = None):
        """随机延迟，模拟人类操作"""
        if min_sec is None:
            min_sec = Config.MIN_DELAY
        if max_sec is None:
            max_sec = Config.MAX_DELAY
        time.sleep(random.uniform(min_sec, max_sec))

    def close(self):
        """关闭浏览器（会话数据已保存到磁盘，下次启动自动恢复）"""
        try:
            if self.context:
                self.context.close()
            if self.playwright:
                self.playwright.stop()
            logger.info("✅ 浏览器已关闭（登录状态已保存，下次启动无需重新登录）")
        except:
            pass

    # ----------------------------------------------------------
    # 主流程
    # ----------------------------------------------------------
    def run(self, keyword: str = None):
        """运行完整采集流程
        
        Args:
            keyword: 搜索关键词
        """
        if keyword:
            Config.KEYWORD = keyword
        
        try:
            # 1. 启动浏览器
            self.init_browser()
            
            # 2. 搜索产品（会等待用户登录）
            self.search_products(Config.KEYWORD)
            
            # 3. 采集列表
            self.collect_product_list()
            
            # 4. 采集详情
            self.collect_product_details()
            
            # 5. 导出数据
            json_path = self.export_to_json()
            excel_path = self.export_to_excel()
            
            # 6. 输出统计
            logger.info("=" * 60)
            logger.info(f"🎉 采集完成!")
            logger.info(f"   采集产品: {len(self.products)} 个")
            logger.info(f"   采集详情: {min(Config.DETAIL_TARGET_COUNT, len(self.products))} 个")
            logger.info(f"   JSON文件: {json_path}")
            if excel_path:
                logger.info(f"   Excel文件: {excel_path}")
            logger.info("=" * 60)
            
            return {
                "status": "success",
                "total": len(self.products),
                "json_path": str(json_path),
                "excel_path": str(excel_path) if excel_path else None
            }
            
        except Exception as e:
            logger.error(f"❌ 采集失败: {e}")
            import traceback
            traceback.print_exc()
            return {"status": "error", "message": str(e)}


# ============================================================
# 入口
# ============================================================
if __name__ == "__main__":
    import sys
    
    # 解析参数
    keyword = "Clitoral Vibrators"
    
    for arg in sys.argv[1:]:
        if arg in ("--help", "-h"):
            print("使用方法: python amazon_scraper.py [关键词]")
            print("  关键词:  搜索关键词，默认 'Clitoral Vibrators'")
            print()
            print("使用说明:")
            print("  1. 运行脚本后会打开Chrome浏览器窗口")
            print("  2. 在浏览器中登录您的亚马逊账号")
            print("  3. 脚本会自动检测登录状态并开始采集")
            sys.exit(0)
        else:
            keyword = arg
    
    print("=" * 60)
    print("  亚马逊产品数据爬虫")
    print("=" * 60)
    print(f"  搜索关键词: {keyword}")
    print(f"  目标数量:   {Config.TARGET_COUNT} 个")
    print(f"  详情数量:   {Config.DETAIL_TARGET_COUNT} 个（含评论）")
    print(f"  评论页数:   {Config.MAX_REVIEW_PAGES} 页/产品")
    print(f"  输出目录:   {Config.OUTPUT_DIR}")
    print("=" * 60)
    print()
    print("📌 操作步骤:")
    print("   1. 脚本会先检查是否有Chrome调试端口可用")
    print("   2. 如果没有，会引导您手动关闭Chrome后重新启动")
    print("   3. Chrome启动后，请在浏览器中手动登录您的账号：")
    print("      • 谷歌账号（可选，保留登录状态）")
    print("      • 亚马逊账号（15306176642）")
    print("   4. 登录完成后，回终端按回车继续")
    print("   5. 采集完成后，浏览器保持打开，下次可直接复用登录状态")
    print()
    
    scraper = AmazonScraper()
    result = scraper.run(keyword)
    
    if result["status"] == "success":
        print(f"\n✅ 采集成功！共 {result['total']} 个产品")
    else:
        print(f"\n❌ 采集失败: {result.get('message', '未知错误')}")